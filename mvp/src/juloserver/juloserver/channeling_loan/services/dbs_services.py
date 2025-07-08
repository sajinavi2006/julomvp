import json
import io
import logging
import openpyxl

from dataclasses import dataclass
from datetime import date, datetime
from typing import List, Any, Dict, Optional, Tuple, Union

from django.conf import settings
from django.db.models import Sum, QuerySet
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.renderers import JSONRenderer

from juloserver.account.models import AccountLimit
from juloserver.application_form.models import OcrKtpResult
from juloserver.channeling_loan.clients import get_dbs_channeling_client
from juloserver.channeling_loan.exceptions import (
    DBSApiError,
    ChannelingLoanStatusNotFound,
    DBSChannelingMappingExcludeJob,
    ChannelingMappingValueError,
    ChannelingMappingNullValueError,
)
from juloserver.channeling_loan.serializers import (
    DBSUpdateLoanStatusRequestSerializer,
    DBSUpdateLoanStatusFailedResponseSerializer,
    DBSUpdateLoanStatusSuccessResponseSerializer,
)
from juloserver.channeling_loan.services.channeling_services import (
    ChannelingMappingServices,
)
from juloserver.channeling_loan.services.general_services import (
    get_credit_score_conversion
)
from juloserver.channeling_loan.constants import (
    ChannelingConst,
    MartialStatusConst,
    ChannelingActionTypeConst,
    ChannelingStatusConst,
    ChannelingLoanRepaymentFileConst,
    ChannelingLoanReconciliationFileConst,
)
from juloserver.channeling_loan.constants.dbs_constants import (
    DBSDisbursementConst,
    DBSApplicationTypeConst,
    DBSEducationConst,
    DBSRepaymentConst,
    DBSChannelingUpdateLoanStatusConst,
    DBSReconciliationConst,
    DBSDisbursementDocumentMappingConst,
    DBSDisbursementDocumentConst,
    DBS_API_CONTENT_TYPE,
)
from juloserver.channeling_loan.models import (
    ChannelingLoanPayment,
    ChannelingLoanStatus,
    DBSChannelingApplicationJob,
)
from juloserver.channeling_loan.forms import (
    RepaymentFileForm,
    ReconciliationFileForm,
)
from juloserver.channeling_loan.tasks import encrypt_data_and_upload_to_dbs_sftp_server
from juloserver.channeling_loan.services.general_services import (
    get_channeling_loan_configuration,
    get_next_filename_counter_suffix,
    create_channeling_loan_send_file_tracking,
    get_channeling_loan_status_by_loan_xid,
    get_channeling_loan_status,
    check_common_failed_channeling,
    update_channeling_loan_status,
    upload_channeling_file_to_oss_and_slack,
    approve_loan_for_channeling,
    create_channeling_loan_api_log,
    # record_channeling_tenure_cap,   # will update when tenure cap release
)
from juloserver.channeling_loan.utils import (
    convert_datetime_to_string,
    GPGTool,
    decrypt_content_with_gpg,
)
from juloserver.julo.constants import RedisLockKeyName
from juloserver.julo.context_managers import redis_lock_for_update
from juloserver.julo.models import Loan
from juloserver.julo.statuses import LoanStatusCodes
from juloserver.julo.clients import get_julo_sentry_client
from juloserver.pii_vault.constants import PiiSource
from juloserver.pii_vault.services import detokenize_for_model_object
from juloserver.sdk.services import xls_to_dict
from juloserver.followthemoney.models import (
    LenderCurrent,
    LoanLenderHistory,
)

from juloserver.loan.tasks.lender_related import loan_lender_approval_process_task

sentry_client = get_julo_sentry_client()
logger = logging.getLogger(__name__)


def process_dbs_file_encryption_and_sftp_upload(
    channeling_type: str,
    action_type: str,
    current_ts: datetime,
    filename_format: str,
    filename_date_format: str,
    content: Any,
    sftp_folder_name: str,
    document_type: str,
    is_upload_to_oss: bool,
) -> str:
    filename_counter_suffix = get_next_filename_counter_suffix(
        channeling_type=channeling_type,
        action_type=action_type,
        current_ts=current_ts,
    )

    filename = filename_format.format(
        current_ts.strftime(filename_date_format), filename_counter_suffix
    )

    encrypt_data_and_upload_to_dbs_sftp_server.delay(content, sftp_folder_name + filename + ".pgp")

    create_channeling_loan_send_file_tracking(
        channeling_type=channeling_type,
        action_type=action_type,
    )

    if is_upload_to_oss:
        lender = LenderCurrent.objects.get_or_none(lender_name=ChannelingConst.LENDER_DBS)
        document_remote_filepath = "channeling_loan/lender_{}/{}".format(lender.id, filename)

        upload_channeling_file_to_oss_and_slack(
            content=content,
            document_remote_filepath=document_remote_filepath,
            lender_id=lender.id,
            filename=filename,
            document_type=document_type,
            channeling_type=channeling_type,
            channeling_action_type=action_type,
            slack_channel=settings.DBS_SLACK_NOTIFICATION_CHANNEL,
        )

    return filename


def encrypt_and_sign_dbs_data_with_gpg(data: str) -> Optional[str]:
    is_success, encrypted_data = GPGTool().encrypt_and_sign(
        content=data,
        recipient=settings.DBS_GPG_ENCRYPT_API_RECIPIENT,
        key_data=settings.DBS_GPG_ENCRYPT_API_KEY_DATA,
        signer_recipient=settings.DBS_GPG_DECRYPT_RECIPIENT,
        signer_key_data=settings.DBS_GPG_DECRYPT_KEY_DATA,
        signer_passphrase=settings.DBS_GPG_DECRYPT_PASSPHRASE,
        custom_compress_algo=DBSDisbursementConst.GPG_ENCRYPT_COMPRESS_ALGORITHM,
    )
    if not is_success:
        logger.error(
            {
                'module': 'juloserver.channeling_loan',
                'action': 'services.dbs_services.encrypt_and_sign_dbs_data_with_gpg',
                'message': encrypted_data,
            }
        )
        sentry_client.captureMessage(
            {
                'error': 'encrypt and sign data with gpg failed',
                'stderr': encrypted_data,
            }
        )
        return None

    return encrypted_data


def decrypt_dbs_data_with_gpg(data: str) -> Optional[str]:
    is_success, decrypted_data = decrypt_content_with_gpg(
        content=data,
        passphrase=settings.DBS_GPG_DECRYPT_PASSPHRASE,
        gpg_recipient=settings.DBS_GPG_DECRYPT_RECIPIENT,
        gpg_key_data=settings.DBS_GPG_DECRYPT_KEY_DATA,
    )
    if not is_success:
        logger.error(
            {
                'module': 'juloserver.channeling_loan',
                'action': 'services.dbs_services.decrypt_dbs_data_with_gpg',
                'message': decrypted_data,
            }
        )
        sentry_client.captureMessage(
            {
                'error': 'decrypt DBS data with gpg failed',
                'stderr': decrypted_data,
            }
        )
        return None

    return decrypted_data


class DBSMappingServices(ChannelingMappingServices):
    @staticmethod
    def get_time_stamp(current_ts: datetime) -> str:
        return convert_datetime_to_string(current_ts)

    @staticmethod
    def get_gender(gender: str) -> int:
        """
        1 - MALE
        2 - FEMALE
        """
        if gender == "Wanita":
            return 2

        return 1

    @staticmethod
    def get_application_type(loan: Loan) -> str:
        loans = Loan.objects.filter(
            customer=loan.customer, loan_status__gte=LoanStatusCodes.CURRENT
        ).exclude(pk=loan.id)

        is_ftc = not loans.exists()
        if is_ftc:
            return DBSApplicationTypeConst.NEW

        return DBSApplicationTypeConst.EXISTING

    @staticmethod
    def get_marital_status(marital_status: str) -> int:
        """
        1 - SINGLE
        2 - DIVORCED
        3 - MARRIED
        """
        if (
            marital_status.upper() == MartialStatusConst.MARRIED.upper()
            or marital_status.upper() == MartialStatusConst.MENIKAH.upper()
        ):
            return 3
        elif marital_status.upper() in [x.upper() for x in MartialStatusConst.DIVORCE_LIST]:
            return 2

        return 1

    @staticmethod
    def get_education_level(last_education) -> int:
        """
        1 - BACHELOR / MASTER / DOCTORAL
        2 - DIPLOMA
        3 - VOCATIONAL
        4 - HIGH SCHOOL
        5 - JUNIOR HIGH SCHOOL
        6 - ELEMENTARY
        7 - NONE
        """
        return DBSEducationConst.LIST.get(last_education.upper(), DBSEducationConst.NOT_FOUND)

    @staticmethod
    def get_dbs_channeling_application_job(customer) -> Optional[DBSChannelingApplicationJob]:
        dbs_channeling_application_job = DBSChannelingApplicationJob.objects.filter(
            job_industry=customer.job_industry, job_description=customer.job_description
        ).last()
        if not dbs_channeling_application_job:
            return None

        if dbs_channeling_application_job.is_exclude:
            raise DBSChannelingMappingExcludeJob(
                'Job is excluded, do not send to DBS',
                mapping_key='POSITION_CODE & NATURE_OF_BUSINESS & AML_RISK_RATING',
                mapping_value='customer.job_industry & customer.job_description',
            )

        return dbs_channeling_application_job

    def get_job_code(self, customer) -> str:
        dbs_channeling_application_job = self.get_dbs_channeling_application_job(customer=customer)
        return dbs_channeling_application_job.job_code if dbs_channeling_application_job else '24'

    def get_job_industry(self, customer) -> str:
        dbs_channeling_application_job = self.get_dbs_channeling_application_job(customer=customer)
        return (
            dbs_channeling_application_job.job_industry_code
            if dbs_channeling_application_job
            else '10'
        )

    def get_aml_risk_rating(self, customer) -> str:
        """1 = High, 2 = Medium, 3 = Low"""
        aml_risk_rating_mapping = {
            'High': '1',
            'Medium': '2',
            'Low': '3',
        }
        dbs_channeling_application_job = self.get_dbs_channeling_application_job(customer=customer)
        return (
            aml_risk_rating_mapping.get(dbs_channeling_application_job.aml_risk_rating)
            if dbs_channeling_application_job
            else '3'
        )

    @staticmethod
    def get_district_code(rt_rw: Optional[str]) -> Optional[str]:
        # District Code is RW, and Sub District Code is RT
        if not rt_rw:
            return None

        _, rw = rt_rw.split("/")

        if len(rw) > 2:
            return rw[-2:]
        return rw

    @staticmethod
    def get_sub_district_code(rt_rw: Optional[str]) -> Optional[str]:
        # District Code is RW, and Sub District Code is RT
        if not rt_rw:
            return None

        rt, _ = rt_rw.split("/")
        return rt

    @staticmethod
    def get_employment_status(employment_status: str) -> str:
        """
        P - PERMANENT
        S - SELF EMPLOYED
        C - CONTRACTUAL
        H - HOUSEWIFE
        O - OTHERS
        """
        # TODO map employment status
        return "P"

    @staticmethod
    def get_yearly_interest_rate(interest_rate_monthly: float) -> int:
        interest_rate = int(round(interest_rate_monthly * 12 * 100, 0))

        # Round to 2 decimal places
        rounded_interest_rate = round(interest_rate, 2)

        # Split into whole and decimal parts
        whole_part = int(rounded_interest_rate)
        decimal_part = int(round((rounded_interest_rate % 1) * 100))

        # Format with leading zeros if necessary
        return int(f"{whole_part:02d}{decimal_part:02d}")

    @staticmethod
    def get_bank_installment_amount(payments: QuerySet) -> int:
        return (
            ChannelingLoanPayment.objects.filter(
                channeling_type=ChannelingConst.DBS,
                payment__in=payments,
            )
            .last()
            .due_amount
        )

    @staticmethod
    def get_total_bank_interest_amount(payments: QuerySet) -> int:
        return ChannelingLoanPayment.objects.filter(
            channeling_type=ChannelingConst.DBS,
            payment__in=payments,
        ).aggregate(Sum('interest_amount'))['interest_amount__sum']

    @staticmethod
    def get_bank_yearly_interest_rate(channeling_loan_config: dict) -> int:
        interest_rate = channeling_loan_config["general"]["INTEREST_PERCENTAGE"]

        # Round to 2 decimal places
        rounded_interest_rate = round(interest_rate, 2)

        # Split into whole and decimal parts
        whole_part = int(rounded_interest_rate)
        decimal_part = int(round((rounded_interest_rate % 1) * 100))

        # Format with leading zeros if necessary
        return int(f"{whole_part:02d}{decimal_part:02d}")

    @staticmethod
    def get_employment_period_in_years(job_start: date) -> str:
        current_ts = timezone.localtime(timezone.now())
        total_employment_month = (current_ts.year - job_start.year) * 12 + (
            current_ts.month - job_start.month
        )
        return str(total_employment_month // 12).zfill(2)

    @staticmethod
    def get_employment_period_in_months(job_start) -> str:
        current_ts = timezone.localtime(timezone.now())
        total_employment_month = (current_ts.year - job_start.year) * 12 + (
            current_ts.month - job_start.month
        )
        return str(total_employment_month % 12).zfill(2)

    @staticmethod
    def get_contract_code(loan_xid: int) -> str:
        return f'{DBSDisbursementConst.PREFIX_CONTRACT_CODE}{str(loan_xid).zfill(12)}'

    @staticmethod
    def get_seller_type(loan_duration: int) -> str:
        if loan_duration in [1, 2, 3, 4]:
            return "NEW"
        return "EXISTING"

    @staticmethod
    def get_customer_credit_score(customer_id) -> Optional[str]:
        return get_credit_score_conversion(customer_id, ChannelingConst.DBS)

    @staticmethod
    def get_employer_phone(customer) -> str:
        employer_phone = customer.company_phone_number
        if not employer_phone:
            application = customer.account.get_active_application()
            employer_phone = application.mobile_phone_1
        return employer_phone

    @staticmethod
    def get_employer_address_street_num(customer) -> str:
        employer_address_street_num = customer.address_street_num
        if not employer_address_street_num:
            application = customer.account.get_active_application()
            employer_address_street_num = application.address_street_num
        return employer_address_street_num


class DBSDisbursementServices:
    def __init__(self):
        self.current_ts = timezone.localtime(timezone.now())
        self.channeling_type = ChannelingConst.DBS
        self.channeling_loan_config = get_channeling_loan_configuration(
            channeling_type=self.channeling_type
        )
        self.module = 'juloserver.channeling_loan'
        self.prefix_action = 'services.dbs_services.DBSDisbursementServices'

    @staticmethod
    def __construct_loan_schedule_list(payments: QuerySet) -> List[Dict[str, Any]]:
        loan_schedules = []
        for payment in payments:
            loan_schedules.append(
                DBSMappingServices(
                    payment=payment,
                    loan=payment.loan,
                    channeling_loan_payment=payment.channelingloanpayment_set.filter(
                        channeling_type=ChannelingConst.DBS,
                    ).last(),
                ).data_mapping(data_map=DBSDisbursementConst.LOAN_SCHEDULE_ELEMENT_MAPPING)
            )
        return loan_schedules

    def construct_disbursement_request_body(self, loan: Loan, x_dbs_uuid: str) -> Dict[str, Any]:
        application = loan.get_application
        customer = loan.customer
        detokenize_customers = detokenize_for_model_object(
            PiiSource.CUSTOMER,
            [
                {
                    'object': customer,
                }
            ],
            force_get_local_data=True,
        )
        payments = loan.payment_set.order_by("payment_number")
        ocr_ktp_result = OcrKtpResult.objects.filter(application_id=application.id).last()

        request_body = DBSMappingServices(
            current_ts=self.current_ts,
            channeling_loan_config=self.channeling_loan_config,
            loan=loan,
            application=application,
            customer=customer,
            detokenize_customer=detokenize_customers[0],
            payments=payments,
            first_payment=payments.first(),
            last_payment=payments.last(),
            rt_rw=ocr_ktp_result.rt_rw if ocr_ktp_result else None,
            job_start=customer.job_start or self.current_ts,
            account_limit=AccountLimit.objects.filter(account=customer.account).last(),
        ).data_mapping(data_map=DBSDisbursementConst.REQUEST_BODY_DATA_MAPPING)

        request_body["header"]["msgId"] = x_dbs_uuid
        request_body["data"]["loanApplicationRequest"][
            "loanScheduleList"
        ] = self.__construct_loan_schedule_list(payments=payments)

        return request_body

    def generate_x_dbs_uuid(self, loan_xid: int) -> str:
        return f'{convert_datetime_to_string(self.current_ts, str_format="%Y%m%d")}{str(loan_xid)}'

    def process_sending_loan_to_dbs(
        self, loan: Loan
    ) -> Union[None, Tuple[bool, bool, str], Tuple[bool, bool, Any]]:
        """
        :param loan:
        :return: a tuple of is_success_hit, should_retry, response
         - is_success_hit: successful hit DBS API or not
         - should_retry: if True, should retry because of request error or encrypt & sign body fail
         - response: response from DBS API
        """
        # CONSTRUCT RAW DISBURSEMENT REQUEST BODY
        x_dbs_uuid = self.generate_x_dbs_uuid(loan_xid=loan.loan_xid)
        try:
            raw_disbursement_request_body = self.construct_disbursement_request_body(
                loan=loan, x_dbs_uuid=x_dbs_uuid
            )
        except ChannelingMappingNullValueError as error:
            # data does not satisfy DBS requirements => return to process in appropriate way
            return False, False, str(error)
        except Exception as error:
            # already automatically captured to Sentry for easily aware
            raise error

        # ENCRYPT AND SIGN REQUEST BODY WITH GPG
        str_raw_disbursement_request_body = json.dumps(raw_disbursement_request_body)
        disbursement_request_body = encrypt_and_sign_dbs_data_with_gpg(
            data=str_raw_disbursement_request_body
        )
        if not disbursement_request_body:
            return False, True, 'Failed to encrypt and sign request body with gpg'

        # SEND LOAN TO DBS
        http_status_code = None
        http_response_text = None
        error_message = None
        output = (False, True, 'Failed to send loan to DBS')
        try:
            response = get_dbs_channeling_client().send_loan(
                loan_id=loan.id,
                x_dbs_uuid=x_dbs_uuid,
                x_dbs_timestamp=convert_datetime_to_string(self.current_ts),
                disbursement_request_body=disbursement_request_body,
            )

            http_status_code = response.status_code
            http_response_text = (
                decrypt_dbs_data_with_gpg(data=response.text) if response.ok else response.text
            )

            output = (True, False, json.loads(http_response_text))

        except (DBSApiError, TypeError, json.JSONDecodeError) as error:
            if isinstance(error, DBSApiError):
                http_status_code = error.response_status_code
                http_response_text = error.response_text
                error_message = error.message

            output = (False, True, 'Failed to send loan to DBS')

            # capture to Sentry for easily aware
            sentry_client.captureException()
        finally:
            create_channeling_loan_api_log(
                channeling_type=self.channeling_type,
                application_id=loan.get_application.id,
                loan_id=loan.id,
                request_type=DBSDisbursementConst.REQUEST_TYPE_API_IN_LOG,
                request=str_raw_disbursement_request_body,
                http_status_code=http_status_code,
                response=http_response_text
                if http_response_text
                else '',  # response in DB is not allow null
                error_message=error_message,
            )

            return output

    @staticmethod
    def parse_error_from_response(channeling_response: Dict[str, Any]) -> Optional[str]:
        error_list = channeling_response.get("data", {}).get("error", {}).get("errorList")
        if error_list:
            return ', '.join([error.get("message", "") for error in error_list])

        error_description = channeling_response.get("error", {}).get("description")
        if error_description:
            return error_description

        validation_infos = channeling_response.get("loanApplicationResponse", {}).get(
            "validationsInfo"
        )
        if validation_infos:
            return ', '.join(
                [validation_info.get("statusDesc", "") for validation_info in validation_infos]
            )

        return None

    def send_loan_for_channeling_to_dbs(self, loan: Loan) -> Tuple[Optional[str], str]:
        def __update_status_and_log(
            status: Optional[str], message: str
        ) -> Tuple[Optional[str], str]:
            if status is not None:
                update_channeling_loan_status(
                    channeling_loan_status_id=channeling_loan_status.id,
                    new_status=status,
                    change_reason=message,
                )

            logger.info(
                {
                    'module': self.module,
                    'action': '{}.send_loan_for_channeling_to_dbs'.format(self.prefix_action),
                    'loan_id': loan.id,
                    'message': message,
                }
            )
            return status, message

        # CHECK DBS CHANNELING FLOW
        target_status = ChannelingStatusConst.PENDING
        lender = LenderCurrent.objects.get_or_none(lender_name=ChannelingConst.LENDER_DBS)
        if lender and lender.is_pre_fund_channeling_flow:
            target_status = ChannelingStatusConst.PREFUND

        # GET CHANNELING LOAN STATUS
        channeling_loan_status = get_channeling_loan_status(loan=loan, status=target_status)
        if not channeling_loan_status:
            raise ChannelingLoanStatusNotFound

        # CHECK COMMON FAILED CHANNELING LOAN STATUS
        reason = check_common_failed_channeling(loan=loan, config=self.channeling_loan_config)
        if reason:
            return __update_status_and_log(status=ChannelingStatusConst.FAILED, message=reason)

        # CONSTRUCT REQUEST DATA & HIT DISBURSEMENT API FOR SENDING LOAN TO DBS
        is_success_hit, should_retry, channeling_response = self.process_sending_loan_to_dbs(
            loan=loan
        )
        if not is_success_hit:
            return __update_status_and_log(
                # if should_retry=True, keep channeling loan status as pending for retry if needed
                status=None if should_retry else ChannelingStatusConst.FAILED,
                message=channeling_response,
            )

        # CHECK RESPONSE AND UPDATE CHANNELING LOAN STATUS
        error = self.parse_error_from_response(channeling_response=channeling_response)
        if error:
            return __update_status_and_log(status=ChannelingStatusConst.FAILED, message=error)
        else:
            return __update_status_and_log(
                status=ChannelingStatusConst.PROCESS,
                message='Success to send loan to DBS, wait for callback',
            )


class DBSUpdateLoanStatusService:
    def __init__(self):
        self.module = 'juloserver.channeling_loan'
        self.prefix_action = 'services.dbs_services.DBSUpdateLoanStatusService'
        self.channeling_type = ChannelingConst.DBS

    @dataclass
    class DBSErrorResponse:
        code: str = DBSChannelingUpdateLoanStatusConst.ERROR_CODE_VALIDATION_ERROR
        message: Optional[str] = None
        more_info: str = DBSChannelingUpdateLoanStatusConst.ERROR_MESSAGE_VALIDATION_ERROR

    @staticmethod
    def generate_error_data(
        x_dbs_uuid: str, errors: List[DBSErrorResponse] = None, error_message: str = None
    ) -> Dict[str, Any]:
        if errors is None:
            errors = []

        if error_message:
            errors.append(DBSUpdateLoanStatusService.DBSErrorResponse(message=error_message))

        return {
            "header": {
                "msgId": x_dbs_uuid,
                "timeStamp": convert_datetime_to_string(timezone.localtime(timezone.now())),
            },
            "data": {
                "error": {
                    "errorList": [
                        {
                            "code": error.code,
                            "message": error.message or "",
                            "moreInfo": error.more_info,
                        }
                        for error in errors
                    ]
                },
            },
        }

    @staticmethod
    def generate_success_data(x_dbs_uuid: str) -> Dict[str, Any]:
        current_ts = convert_datetime_to_string(timezone.localtime(timezone.now()))
        return {
            "header": {
                "msgId": x_dbs_uuid,
                "timeStamp": current_ts,
            },
            "data": {"loanApplicationStatusResponse": {"receiptDateTime": current_ts}},
        }

    def generate_serializer_errors(
        self, x_dbs_uuid: str, request_serializer_errors: Dict[str, Any]
    ) -> Dict[str, Any]:
        return self.generate_error_data(
            x_dbs_uuid=x_dbs_uuid,
            errors=[
                DBSUpdateLoanStatusService.DBSErrorResponse(
                    code=DBSChannelingUpdateLoanStatusConst.ERROR_CODE_INVALID_INPUT_PARAM,
                    message=error_field,
                    more_info=DBSChannelingUpdateLoanStatusConst.ERROR_MESSAGE_INVALID_INPUT_PARAM,
                )
                for error_field, list_message in request_serializer_errors.items()
            ],
        )

    @staticmethod
    def get_headers(request_metas: Dict[str, str]) -> Tuple[bool, Union[Dict[str, Any], str]]:
        required_headers = DBSChannelingUpdateLoanStatusConst.HTTP_REQUIRED_HEADERS
        optional_headers = DBSChannelingUpdateLoanStatusConst.HTTP_OPTIONAL_HEADERS

        missing_headers = [header for header in required_headers if header not in request_metas]
        if missing_headers:
            return False, f'Missing required headers: {", ".join(missing_headers)}'

        header_values = {
            header: request_metas.get(header) for header in required_headers + optional_headers
        }
        return True, header_values

    @staticmethod
    def get_channeling_loan_status(
        loan_xid: str,
    ) -> Tuple[bool, Union[ChannelingLoanStatus, str]]:
        channeling_loan_status = get_channeling_loan_status_by_loan_xid(
            loan_xid=loan_xid, channeling_type=ChannelingConst.DBS
        )
        if channeling_loan_status is None:
            return False, "applicationId not found"
        elif channeling_loan_status.channeling_status != ChannelingStatusConst.PROCESS:
            return False, "applicationId already be processed"

        return True, channeling_loan_status

    def response(
        self,
        is_error: bool,
        x_dbs_uuid: str,
        data: Union[str, Dict[str, Any]] = None,
        loan: Optional[Loan] = None,
        decrypted_request_body: Optional[str] = None,
    ) -> HttpResponse:
        if is_error:
            if isinstance(data, str):
                data = self.generate_error_data(x_dbs_uuid=x_dbs_uuid, error_message=data)
            else:
                data = self.generate_serializer_errors(
                    x_dbs_uuid=x_dbs_uuid, request_serializer_errors=data
                )
        else:
            data = self.generate_success_data(x_dbs_uuid=x_dbs_uuid)

        if is_error:
            status_code = status.HTTP_400_BAD_REQUEST
            serializer = DBSUpdateLoanStatusFailedResponseSerializer(data)
        else:
            status_code = status.HTTP_200_OK
            serializer = DBSUpdateLoanStatusSuccessResponseSerializer(data)

        response_text = JSONRenderer().render(serializer.data).decode()

        create_channeling_loan_api_log(
            channeling_type=self.channeling_type,
            application_id=loan.get_application.id if loan else None,
            loan_id=loan.id if loan else None,
            request_type=DBSDisbursementConst.REQUEST_TYPE_CALLBACK_IN_LOG,
            request=decrypted_request_body,
            http_status_code=status_code,
            response=response_text,
            error_message=str(data) if is_error else None,
        )

        return HttpResponse(
            encrypt_and_sign_dbs_data_with_gpg(data=response_text),
            status=status_code,
            content_type=DBS_API_CONTENT_TYPE,
        )

    def process_request(self, request):
        logger.info(
            {
                'module': self.module,
                'action': '{}.process_request'.format(self.prefix_action),
                'header': request.META,
                'request_data': request.body,
            }
        )

        # CHECK HEADER
        is_valid_header, dbs_headers = self.get_headers(request_metas=request.META)
        if not is_valid_header:
            return self.response(is_error=True, x_dbs_uuid='', data=dbs_headers)

        x_dbs_uuid = dbs_headers[DBSChannelingUpdateLoanStatusConst.HTTP_X_DBS_UUID]

        # DECRYPT REQUEST BODY
        decrypted_request_body = decrypt_dbs_data_with_gpg(data=request.body.decode('utf-8'))
        try:
            request_body = json.loads(decrypted_request_body)
        except (TypeError, json.JSONDecodeError):
            return self.response(
                is_error=True,
                x_dbs_uuid=x_dbs_uuid,
                data='Failed to decrypt and parse request body to json',
                decrypted_request_body=decrypted_request_body,
            )

        # CHECK REQUEST BODY
        request_serializer = DBSUpdateLoanStatusRequestSerializer(
            data=request_body, context=dbs_headers
        )
        if not request_serializer.is_valid():
            return self.response(
                is_error=True,
                x_dbs_uuid=x_dbs_uuid,
                data=request_serializer.errors,
                decrypted_request_body=decrypted_request_body,
            )

        # CHECK CONTRACT NUMBER
        loan_xid = request_serializer.loan_xid
        is_valid_loan_xid, channeling_loan_status = self.get_channeling_loan_status(
            loan_xid=loan_xid
        )
        if not is_valid_loan_xid:
            return self.response(
                is_error=True,
                x_dbs_uuid=x_dbs_uuid,
                data=channeling_loan_status,
                decrypted_request_body=decrypted_request_body,
            )

        loan = channeling_loan_status.loan
        # LOCK TO AVOID DUPLICATED CALLBACK & PROCESS APPROVAL CHANNELING LOAN => CALL ASYNC TASK
        loan_id = channeling_loan_status.loan_id  # noqa
        with redis_lock_for_update(RedisLockKeyName.APPROVE_LOAN_FOR_CHANNELING, loan_id):
            is_rejected = request_serializer.is_rejected

            if is_rejected:
                approval_status = 'n'
            else:
                approval_status = 'y'
                # will update when tenure cap release
                # if loan.loan_status == LoanStatusCodes.CURRENT:
                #     record_channeling_tenure_cap(
                #         loan.loan_duration,
                #         loan.loan_amount,
                #         ChannelingConst.DBS,
                #     )

            approve_loan_for_channeling(
                loan=channeling_loan_status.loan,
                channeling_type=ChannelingConst.DBS,
                approval_status=approval_status,
                reason=None if not is_rejected else request_serializer.rejected_reason,
            )
            if loan.status == LoanStatusCodes.LENDER_APPROVAL:
                lender = LenderCurrent.objects.get_or_none(lender_name=ChannelingConst.LENDER_DBS)
                if lender and lender.is_pre_fund_channeling_flow:
                    self.continue_process_to_next_loan_status(loan, approval_status)

        return self.response(
            x_dbs_uuid=x_dbs_uuid,
            is_error=False,
            loan=channeling_loan_status.loan,
            decrypted_request_body=decrypted_request_body,
        )

    def continue_process_to_next_loan_status(self, loan, approval_status):
        lender_id = loan.lender_id
        if approval_status == 'n':
            LoanLenderHistory.objects.create(loan=loan, lender_id=lender_id)
            loan.update_safely(lender=None)
        loan_lender_approval_process_task.delay(loan.id, lender_ids=[lender_id])


class DBSRepaymentServices:
    def __init__(self):
        self.action_type = ChannelingActionTypeConst.REPAYMENT
        self.channeling_type = ChannelingConst.DBS
        self.current_ts = timezone.localtime(timezone.now())
        self.module = 'juloserver.channeling_loan'
        self.prefix_action = 'services.dbs_services.DBSRepaymentServices'

    def get_dbs_repayment_data(self, request) -> List[Dict[str, Any]]:
        """
        Convert excel to List[Dict[str, Any]] format so can reuse main function
        capture error in case conversion fail
        """
        upload_form = RepaymentFileForm(request.POST, request.FILES)
        if not upload_form.is_valid():
            raise Exception("Invalid form")

        file_ = upload_form.cleaned_data["repayment_file_field"]
        extension = file_.name.split(".")[-1]

        if extension not in ChannelingConst.FILE_UPLOAD_EXTENSIONS:
            raise Exception("Please upload correct file excel")

        try:
            list_data = []
            excel_datas = xls_to_dict(file_)

            for sheet in excel_datas:
                for _, row in enumerate(excel_datas[sheet]):
                    list_data.append(row)

            return list_data
        except Exception as error:
            logger.exception(
                {
                    'module': self.module,
                    'action': '{}.get_dbs_repayment_data'.format(self.prefix_action),
                    'message': 'Failed to construct channeling repayment data',
                    'error': error,
                }
            )
            raise Exception(error)

    @staticmethod
    def construct_dbs_repayment_request_xlsx_bytes(
        headers: Dict[str, Any],
        sheet_name: str,
        list_data: List[Dict[str, Any]],
    ) -> Tuple[bool, Union[bytes, str]]:
        errors = []

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Rename the first sheet
        sheet.title = sheet_name

        # Write header row -> first row
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=column_index, value=headers.get(header))

        # Write data rows -> from second row until end of list_data
        for row_index, key_value in enumerate(list_data, start=2):
            for column_index, header in enumerate(headers, start=1):
                if header not in key_value:
                    errors.append("Row {}: {} is missing".format(column_index, header))
                    continue

                value = key_value.get(header, '')
                if "payment" in header:
                    value = int(key_value.get(header, 0))

                if "loan_xid" in header:
                    value = DBSRepaymentConst.LOAN_ACC_NUMBER_PREFIX + value

                if "date" in header:
                    try:
                        datetime.strptime(key_value[header], "%d/%m/%Y")
                    except ValueError:
                        errors.append("Row {}: {} is not a valid date".format(column_index, header))
                        continue

                sheet.cell(row=row_index, column=column_index, value=value)

        # Save the workbook to a BytesIO object
        excel_byte_array = io.BytesIO()
        workbook.save(excel_byte_array)

        if errors:
            return False, '\n'.join(errors)
        return True, excel_byte_array.getvalue()

    def send_repayment_for_channeling_to_dbs(
        self, request, is_upload_to_oss=False
    ) -> Tuple[bool, str]:
        """
        main function to process the form request
        form are validated, then sent to DBS SFTP
        """
        if self.channeling_type != ChannelingConst.DBS:
            return False, 'Wrong channeling type'

        try:
            list_repayment_data = self.get_dbs_repayment_data(request)
        except Exception as error:
            return False, str(error)

        is_success, content = self.construct_dbs_repayment_request_xlsx_bytes(
            headers=DBSRepaymentConst.REPAYMENT_DATA_HEADER_DICTIONARY,
            sheet_name=DBSRepaymentConst.SHEET_NAME,
            list_data=list_repayment_data,
        )

        if not is_success:
            return False, content

        filename = process_dbs_file_encryption_and_sftp_upload(
            channeling_type=self.channeling_type,
            action_type=self.action_type,
            current_ts=self.current_ts,
            filename_format=DBSRepaymentConst.FILENAME_FORMAT,
            filename_date_format=DBSRepaymentConst.FILENAME_DATE_FORMAT,
            content=content,
            sftp_folder_name=DBSRepaymentConst.REQUEST_FOLDER_NAME,
            document_type=ChannelingLoanRepaymentFileConst.DOCUMENT_TYPE,
            is_upload_to_oss=is_upload_to_oss,
        )

        logger.info(
            {
                'module': self.module,
                'action': '{}.send_repayment_for_channeling_to_dbs'.format(self.prefix_action),
                'number_of_payments': len(list_repayment_data),
                'filename': filename,
                'message': 'Send repayment for DBS channeling successful',
            }
        )

        return is_success, filename


class DBSReconciliationServices:
    def __init__(self):
        self.action_type = ChannelingActionTypeConst.RECONCILIATION
        self.channeling_type = ChannelingConst.DBS
        self.current_ts = timezone.localtime(timezone.now())
        self.module = 'juloserver.channeling_loan'
        self.prefix_action = 'services.dbs_services.DBSReconciliationServices'

    def get_dbs_reconciliation_data(self, request) -> List[Dict[str, Any]]:
        """
        Convert excel to List[Dict[str, Any]] format so can reuse main function
        capture error in case conversion fail
        """
        upload_form = ReconciliationFileForm(request.POST, request.FILES)
        if not upload_form.is_valid():
            raise Exception("Invalid form")

        file_ = upload_form.cleaned_data["reconciliation_file_field"]
        extension = file_.name.split(".")[-1]

        if extension not in ChannelingConst.FILE_UPLOAD_EXTENSIONS:
            raise Exception("Please upload correct file excel")

        try:
            list_data = []
            excel_datas = xls_to_dict(file_)

            for sheet in excel_datas:
                for _, row in enumerate(excel_datas[sheet]):
                    list_data.append(row)

            return list_data
        except Exception as error:
            logger.exception(
                {
                    'module': self.module,
                    'action': '{}.get_dbs_reconciliation_data'.format(self.prefix_action),
                    'message': 'Failed to construct channeling reconciliation data',
                    'error': error,
                }
            )
            raise Exception(error)

    @staticmethod
    def construct_dbs_reconciliation_request_txt_content(
        dict_rows: List[dict], data_header_list: List[str]
    ) -> Tuple[bool, str]:
        errors = []
        result_rows = []

        for index, dict_row in enumerate(dict_rows):
            elements = []
            for field in data_header_list:
                if field not in dict_row:
                    errors.append("Row {}: {} is missing".format(index + 1, field))
                    continue
                if "date" in field:
                    try:
                        datetime.strptime(dict_row[field], "%d/%m/%Y")
                    except ValueError:
                        errors.append("Row {}: {} is not a valid date".format(index + 1, field))
                        continue
                element = dict_row[field]
                elements.append(element + "|")
            result_rows.append("".join(elements))

        if errors:
            return False, '\n'.join(errors)
        return True, '\n'.join(result_rows)

    def send_reconciliation_for_channeling_to_dbs(
        self, request, is_upload_to_oss=False
    ) -> Tuple[bool, str]:
        if self.channeling_type != ChannelingConst.DBS:
            return False, 'Wrong channeling type'

        try:
            list_reconciliation_data = self.get_dbs_reconciliation_data(request)
        except Exception as error:
            return False, str(error)

        is_success, content = self.construct_dbs_reconciliation_request_txt_content(
            dict_rows=list_reconciliation_data,
            data_header_list=DBSReconciliationConst.RECONCILIATION_DATA_HEADER_LIST,
        )

        if not is_success:
            return False, content

        filename = process_dbs_file_encryption_and_sftp_upload(
            channeling_type=self.channeling_type,
            action_type=self.action_type,
            current_ts=self.current_ts,
            filename_format=DBSReconciliationConst.FILENAME_FORMAT,
            filename_date_format=DBSReconciliationConst.FILENAME_DATE_FORMAT,
            content=content,
            sftp_folder_name=DBSReconciliationConst.REQUEST_FOLDER_NAME,
            document_type=ChannelingLoanReconciliationFileConst.DOCUMENT_TYPE,
            is_upload_to_oss=is_upload_to_oss,
        )

        logger.info(
            {
                'module': self.module,
                'action': '{}.send_reconciliation_for_channeling_to_dbs'.format(self.prefix_action),
                'number_of_reconciliations': len(list_reconciliation_data),
                'filename': filename,
                'message': 'Send reconciliation for DBS channeling successful',
            }
        )

        return is_success, filename


class DBSDisbursementDocumentMappingServices(ChannelingMappingServices):
    @staticmethod
    def get_gender(gender: str) -> int:
        """
        1 - MALE
        2 - FEMALE
        """
        if gender == "Wanita":
            return 2

        return 1

    @staticmethod
    def get_education_level(last_education) -> int:
        """
        1 - BACHELOR / MASTER / DOCTORAL
        2 - DIPLOMA
        3 - VOCATIONAL
        4 - HIGH SCHOOL
        5 - JUNIOR HIGH SCHOOL
        6 - ELEMENTARY
        7 - NONE
        """
        return DBSEducationConst.LIST.get(last_education.upper(), DBSEducationConst.NOT_FOUND)

    @staticmethod
    def get_marital_status(marital_status: str) -> int:
        """
        1 - SINGLE
        2 - DIVORCED
        3 - MARRIED
        If Janda/Duda map to 2
        """
        if (
            marital_status.upper() == MartialStatusConst.MARRIED.upper()
            or marital_status.upper() == MartialStatusConst.MENIKAH.upper()
        ):
            return 3
        elif marital_status.upper() in [x.upper() for x in MartialStatusConst.DIVORCE_LIST]:
            return 2

        return 1

    @staticmethod
    def get_district_code(rt_rw: Optional[str]) -> Optional[str]:
        # District Code is RW, and Sub District Code is RT
        if not rt_rw:
            return None

        _, rw = rt_rw.split("/")

        if len(rw) > 2:
            return rw[-2:]
        return rw

    @staticmethod
    def get_sub_district_code(rt_rw: Optional[str]) -> Optional[str]:
        # District Code is RW, and Sub District Code is RT
        if not rt_rw:
            return None

        rt, _ = rt_rw.split("/")
        return rt

    @staticmethod
    def get_dbs_channeling_application_job(customer) -> Optional[DBSChannelingApplicationJob]:
        dbs_channeling_application_job = DBSChannelingApplicationJob.objects.filter(
            job_industry=customer.job_industry, job_description=customer.job_description
        ).last()
        if not dbs_channeling_application_job:
            return None

        if dbs_channeling_application_job.is_exclude:
            raise DBSChannelingMappingExcludeJob(
                'Job is excluded, do not send to DBS',
                mapping_key='POSITION_CODE & NATURE_OF_BUSINESS & AML_RISK_RATING',
                mapping_value='customer.job_industry & customer.job_description',
            )

        return dbs_channeling_application_job

    def get_job_code(self, customer) -> str:
        dbs_channeling_application_job = self.get_dbs_channeling_application_job(customer=customer)
        return dbs_channeling_application_job.job_code if dbs_channeling_application_job else '24'

    def get_job_industry(self, customer) -> str:
        dbs_channeling_application_job = self.get_dbs_channeling_application_job(customer=customer)
        return (
            dbs_channeling_application_job.job_industry_code
            if dbs_channeling_application_job
            else '10'
        )

    def get_aml_risk_rating(self, customer) -> str:
        """1 = High, 2 = Medium, 3 = Low"""
        aml_risk_rating_mapping = {
            'High': '1',
            'Medium': '2',
            'Low': '3',
        }
        dbs_channeling_application_job = self.get_dbs_channeling_application_job(customer=customer)
        return (
            aml_risk_rating_mapping.get(dbs_channeling_application_job.aml_risk_rating)
            if dbs_channeling_application_job
            else '3'
        )

    @staticmethod
    def get_employment_period(job_start: date) -> str:
        current_ts = timezone.localtime(timezone.now())
        total_employment_month = (current_ts.year - job_start.year) * 12 + (
            current_ts.month - job_start.month
        )
        total_year = total_employment_month // 12
        total_rest_month = total_employment_month % 12
        return str(total_year).zfill(2) + str(total_rest_month).zfill(2)

    @staticmethod
    def get_contract_code(loan_xid: int) -> str:
        return (
            f'{DBSDisbursementDocumentMappingConst.PREFIX_CONTRACT_CODE}{str(loan_xid).zfill(12)}'
        )

    @staticmethod
    def get_bank_yearly_interest_rate(channeling_loan_config: dict) -> str:
        interest_rate = channeling_loan_config["general"]["INTEREST_PERCENTAGE"]

        # Round to 2 decimal places
        rounded_interest_rate = round(interest_rate, 2)

        # Split into whole and decimal parts
        whole_part = int(rounded_interest_rate)
        decimal_part = int(round((rounded_interest_rate % 1) * 100))

        # Format with leading zeros if necessary
        return f"{whole_part:02d}{decimal_part:02d}"

    @staticmethod
    def get_current_company_on_job_in_years(job_start: Optional[date]) -> str:
        current_ts = timezone.localtime(timezone.now())
        total_employment_month = (current_ts.year - job_start.year) * 12 + (
            current_ts.month - job_start.month
        )
        return str(total_employment_month // 12).zfill(2)

    @staticmethod
    def get_current_company_on_job_in_months(job_start: Optional[date]) -> str:
        current_ts = timezone.localtime(timezone.now())
        total_employment_month = (current_ts.year - job_start.year) * 12 + (
            current_ts.month - job_start.month
        )
        return str(total_employment_month).zfill(2)

    @staticmethod
    def get_total_bank_interest_amount(payments: QuerySet) -> int:
        return ChannelingLoanPayment.objects.filter(
            channeling_type=ChannelingConst.DBS,
            payment__in=payments,
        ).aggregate(Sum('interest_amount'))['interest_amount__sum']

    @staticmethod
    def get_bank_installment_amount(payments: QuerySet) -> int:
        return (
            ChannelingLoanPayment.objects.filter(
                channeling_type=ChannelingConst.DBS,
                payment__in=payments,
            )
            .last()
            .due_amount
        )

    @staticmethod
    def get_yearly_interest_rate(interest_rate_monthly: float) -> str:
        interest_rate = int(round(interest_rate_monthly * 12 * 100, 0))

        # Round to 2 decimal places
        rounded_interest_rate = round(interest_rate, 2)

        # Split into whole and decimal parts
        whole_part = int(rounded_interest_rate)
        decimal_part = int(round((rounded_interest_rate % 1) * 100))

        # Format with leading zeros if necessary
        return f"{whole_part:02d}{decimal_part:02d}"

    @staticmethod
    def get_customer_loan_sequence(loan: Loan) -> int:
        return Loan.objects.filter(
            customer=loan.customer, loan_status__gte=LoanStatusCodes.CURRENT, id__lte=loan.id
        ).count()

    @staticmethod
    def get_installment_amount(channeling_loan_payment: ChannelingLoanPayment) -> int:
        return channeling_loan_payment.principal_amount + channeling_loan_payment.interest_amount


class DBSDisbursementDocumentServices:
    def __init__(self):
        self.action_type = ChannelingActionTypeConst.DISBURSEMENT
        self.current_ts = timezone.localtime(timezone.now())
        self.channeling_type = ChannelingConst.DBS
        self.channeling_loan_config = get_channeling_loan_configuration(
            channeling_type=self.channeling_type
        )
        self.module = 'juloserver.channeling_loan'
        self.prefix_action = 'services.dbs_services.DBSDisbursementDocumentServices.'

    def __init_mapping_service(self, loan: Loan) -> DBSDisbursementDocumentMappingServices:
        application = loan.get_application
        customer = loan.customer
        detokenize_customers = detokenize_for_model_object(
            PiiSource.CUSTOMER,
            [
                {
                    'object': customer,
                }
            ],
            force_get_local_data=True,
        )
        payments = loan.payment_set.order_by("payment_number")

        return DBSDisbursementDocumentMappingServices(
            current_ts=self.current_ts,
            channeling_loan_config=self.channeling_loan_config,
            loan=loan,
            application=application,
            customer=customer,
            detokenize_customer=detokenize_customers[0],
            ocr_ktp_result=OcrKtpResult.objects.filter(application_id=application.id).last(),
            job_start=customer.job_start or self.current_ts,
            payments=payments,
            first_payment=payments.first(),
            last_payment=payments.last(),
            account_limit=AccountLimit.objects.filter(account=customer.account).last(),
        )

    def __map_loan_installment_list(self, loan: Loan) -> List[Dict[str, Any]]:
        loan_installments = []
        payments = loan.payment_set.select_related("loan").order_by("payment_number")
        for payment in payments:
            channeling_loan_payment = payment.channelingloanpayment_set.filter(  # noqa
                channeling_type=self.channeling_type,
            ).last()

            loan_installments.append(
                DBSDisbursementDocumentMappingServices(
                    loan=loan,
                    payment=payment,
                    channeling_loan_payment=channeling_loan_payment,
                ).data_mapping(
                    data_map=DBSDisbursementDocumentMappingConst.INSTALLMENT_ELEMENT_MAPPING
                )
            )
        return loan_installments

    def map_list_disbursement_documents_data(
        self, channeling_loan_statuses: List[ChannelingLoanStatus]
    ) -> Tuple[
        List[int], List[int], Dict[int, str], List[Dict], List[Dict], List[Dict], List[Dict]
    ]:
        """
        :param channeling_loan_statuses: list of channeling loan statuses need to map data
        :return: only return success channeling loan statues and success mapping data, skip failed
        """
        list_application = []
        list_contract = []
        list_goods_info = []
        list_loan_installment = []
        success_mapping_channeling_loan_status_ids = []
        failed_mapping_channeling_loan_status_ids = []
        failed_mapping_channeling_loan_status_errors = {}

        for channeling_loan_status in channeling_loan_statuses:
            loan_id = channeling_loan_status.loan_id  # noqa
            base_logger_data = {
                'module': self.module,
                'action': self.prefix_action + 'construct_list_disbursement_documents_data',
                'loan_id': loan_id,
            }

            try:
                loan = channeling_loan_status.loan
                mapping_service = self.__init_mapping_service(loan=loan)

                list_application.append(
                    mapping_service.data_mapping(
                        data_map=DBSDisbursementDocumentMappingConst.APPLICATION_MAPPING
                    )
                )
                list_contract.append(
                    mapping_service.data_mapping(
                        data_map=DBSDisbursementDocumentMappingConst.CONTRACT_MAPPING
                    )
                )
                list_goods_info.append(
                    mapping_service.data_mapping(
                        data_map=DBSDisbursementDocumentMappingConst.GOODS_INFO_MAPPING
                    )
                )
                list_loan_installment.extend(self.__map_loan_installment_list(loan=loan))

                success_mapping_channeling_loan_status_ids.append(channeling_loan_status.id)
            except Exception as e:
                failed_mapping_channeling_loan_status_ids.append(channeling_loan_status.id)
                failed_mapping_channeling_loan_status_errors[channeling_loan_status.id] = str(e)

                if not isinstance(e, ChannelingMappingValueError):
                    sentry_client.captureException()

                logger.exception(
                    {
                        **base_logger_data,
                        'message': 'Failed to map DBS channeling disbursement document data',
                        'error': e,
                    }
                )
                continue  # skip to the next loan

        return (
            success_mapping_channeling_loan_status_ids,
            failed_mapping_channeling_loan_status_ids,
            failed_mapping_channeling_loan_status_errors,
            list_application,
            list_contract,
            list_goods_info,
            list_loan_installment,
        )

    @staticmethod
    def __construct_document_content(
        rows: List[Dict[str, str]],
        headers: List[str],
        separator: str = DBSDisbursementDocumentConst.CONTENT_SEPARATOR,
    ) -> str:
        """
        Convert a list of dictionaries to separator-delimited text format.
        Example:
            rows = [
                {'column1': '123', 'column2': '234'},
                {'column1': '3', 'column2': '23'}
            ]
            Result is:
            column1|column2
            123|234
            3|23
        :param rows: List of dictionaries where each dictionary represents a row of data
        :param headers: List of headers to use for the first row
        :param separator: separator character to use
        :return: separator-delimited text content with header row
        """
        # Create header row
        result = [separator.join(headers)]

        # Create data rows
        for row in rows:
            row_values = [str(row.get(header, '')) for header in headers]
            result.append(separator.join(row_values))

        # Join all rows with newlines
        return "\n".join(result)

    def __construct_loan_installment_list_content(self, rows: List[Dict[str, str]]) -> str:
        # Create header row with date will be processed in DBS system. Format: DD-MM-YYYY
        result = [
            DBSDisbursementDocumentConst.LOAN_INSTALLMENT_HEADER.format(
                self.current_ts.strftime("%d-%m-%Y")
            )
        ]

        # Create data rows
        column_names = DBSDisbursementDocumentMappingConst.INSTALLMENT_ELEMENT_MAPPING.keys()
        for row in rows:
            row_values = [str(row.get(column_name, '')) for column_name in column_names]
            result.append(
                DBSDisbursementDocumentConst.LOAN_INSTALLMENT_LIST_CONTENT_SEPARATOR.join(
                    row_values
                )
            )

        # Add footer in the end with count of detail records excluding the header and trailer
        result.append(
            DBSDisbursementDocumentConst.LOAN_INSTALLMENT_FOOTER.format(str(len(rows)).zfill(10))
        )

        # Join all rows with newlines
        return "\n".join(result)

    def construct_disbursement_document_files(
        self, channeling_loan_statuses: List[ChannelingLoanStatus]
    ) -> Tuple[List[int], List[int], Dict[int, str], Dict[str, str]]:
        """
        :param channeling_loan_statuses: list of channeling loan statuses need to construct data
        :return: only return success channeling loan statues & success construct data, skip failed.
        A tuple of:
            - List of success channeling loan statuses
            - A dictionary: key is file name contained first format is the datetime,
            second format is the numerical order; and value is content of file:
                + Application
                + Contract
                + Goods Info
                + Loan Installment List
        """
        (
            success_mapping_channeling_loan_status_ids,
            failed_mapping_channeling_loan_status_ids,
            failed_mapping_channeling_loan_status_errors,
            list_application,
            list_contract,
            list_goods_info,
            list_loan_installment,
        ) = self.map_list_disbursement_documents_data(
            channeling_loan_statuses=channeling_loan_statuses
        )

        return (
            success_mapping_channeling_loan_status_ids,
            failed_mapping_channeling_loan_status_ids,
            failed_mapping_channeling_loan_status_errors,
            {
                DBSDisbursementDocumentConst.APPLICATION_FILENAME_FORMAT: (
                    self.__construct_document_content(
                        rows=list_application,
                        headers=DBSDisbursementDocumentMappingConst.APPLICATION_MAPPING.keys(),
                    )
                ),
                DBSDisbursementDocumentConst.CONTRACT_FILENAME_FORMAT: (
                    self.__construct_document_content(
                        rows=list_contract,
                        headers=DBSDisbursementDocumentMappingConst.CONTRACT_MAPPING.keys(),
                    )
                ),
                DBSDisbursementDocumentConst.GOODS_INFO_FILENAME_FORMAT: (
                    self.__construct_document_content(
                        rows=list_goods_info,
                        headers=DBSDisbursementDocumentMappingConst.GOODS_INFO_MAPPING.keys(),
                    )
                ),
                DBSDisbursementDocumentConst.LOAN_INSTALLMENT_LIST_FILENAME_FORMAT: (
                    self.__construct_loan_installment_list_content(rows=list_loan_installment)
                ),
            },
        )
