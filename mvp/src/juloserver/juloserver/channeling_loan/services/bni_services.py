from typing import Dict, Any, List, Tuple, Optional, Iterator, Union
import logging
import openpyxl
import zipfile

from io import BytesIO
from django.utils import timezone

from juloserver.channeling_loan.constants.bni_constants import (
    BNIDisbursementConst,
    BNIRepaymentConst,
    BNISupportingDisbursementDocumentConst,
)

from juloserver.channeling_loan.clients import get_bni_sftp_client
from juloserver.channeling_loan.constants import (
    ChannelingConst,
    ChannelingStatusConst,
    ChannelingActionTypeConst,
    ChannelingApprovalStatusConst,
)

from juloserver.channeling_loan.exceptions import (
    BNIChannelingLoanKTPNotFound,
    BNIChannelingLoanSelfieNotFound,
    BNIChannelingLoanSKRTPNotFound,
)
from juloserver.channeling_loan.forms import (
    RepaymentFileForm,
)
from juloserver.channeling_loan.models import ChannelingLoanStatus
from juloserver.channeling_loan.services.channeling_services import (
    ChannelingMappingServices,
)
from juloserver.channeling_loan.services.general_services import (
    get_next_filename_counter_suffix,
    SFTPProcess,
    create_channeling_loan_send_file_tracking,
    get_channeling_loan_configuration,
    bulk_update_channeling_loan_status,
)

from juloserver.channeling_loan.utils import download_file_from_oss, convert_datetime_to_string
from juloserver.julo.clients import get_julo_sentry_client
from juloserver.julo.models import Customer, Loan, Image, Document
from juloserver.pii_vault.constants import PiiSource
from juloserver.pii_vault.services import detokenize_for_model_object
from juloserver.sdk.services import xls_to_dict


sentry_client = get_julo_sentry_client()
logger = logging.getLogger(__name__)


def construct_bni_xlsx_bytes(
    headers: List[str],
    sheet_name: str,
    list_data: List[Dict[str, Any]],
    header_no: str,
    header_map: Optional[Dict[str, str]],
) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Rename the first sheet
    sheet.title = sheet_name

    # Write header row -> first row
    for column_index, header in enumerate(headers, start=1):
        if header_map and header in header_map:
            header = header_map[header]
        sheet.cell(row=1, column=column_index, value=header)

    # Write data rows -> from second row until end of list_data
    for row_index, data in enumerate(list_data, start=2):
        for column_index, header in enumerate(headers, start=1):
            if header == header_no:
                # This value is hardcode during mapping data, so we need to calculate it
                # e.g. row_index = 3, value NO = 2
                sheet.cell(row=row_index, column=column_index, value=row_index - 1)
            else:
                sheet.cell(row=row_index, column=column_index, value=data.get(header, None))

    # Save the workbook to a BytesIO object
    excel_byte_array = BytesIO()
    workbook.save(excel_byte_array)

    # Get the byte array
    excel_byte_array = excel_byte_array.getvalue()

    return excel_byte_array


def send_file_for_channeling_to_bni(
    data_bytes: bytes,
    folder_name: str,
    filename_format: str,
    filename_datetime_format: str,
    filename_counter_suffix: str,
    current_ts: timezone,
) -> str:
    """
    data_bytes is data (excel / zip) that going to be sent to BNI sftp
    filename_format is format of the filename
    filename_datetime_format is format of the date in filename
    filename_counter_suffix is counter to count how many file have been sent today
    """
    filename = filename_format.format(
        current_ts.strftime(filename_datetime_format),
        filename_counter_suffix,
    )

    # UPLOAD FILE TO BNI SFTP SERVER
    SFTPProcess(sftp_client=get_bni_sftp_client()).upload(
        content=data_bytes, remote_path='{}/{}'.format(folder_name, filename)
    )

    return filename


class BNIMappingServices(ChannelingMappingServices):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    @staticmethod
    def format_tanggal_tagihan(day: int):
        return str(day).zfill(2)

    @staticmethod
    def get_emergency_contact(customer: Customer):
        return customer.spouse_name if customer.spouse_name else customer.kin_name

    @staticmethod
    def get_hubungan(customer: Customer):
        if customer.spouse_name:
            return "Istri" if customer.gender == "Pria" else "Suami"
        return customer.kin_relationship

    @staticmethod
    def get_telp_ec(customer: Customer):
        return customer.spouse_mobile_phone if customer.spouse_name else customer.kin_mobile_phone


class BNIDisbursementServices:
    def __init__(self):
        self.current_ts = timezone.localtime(timezone.now())
        self.channeling_type = ChannelingConst.BNI
        self.module = 'juloserver.channeling_loan'
        self.prefix_action = 'services.bni_services.BNIDisbursementServices'

    @staticmethod
    def construct_bni_disbursement_data(loan: Loan) -> Dict[str, Any]:
        customer = loan.customer
        detokenize_customers = detokenize_for_model_object(
            PiiSource.CUSTOMER,
            [
                {
                    'object': customer,
                }
            ],
            force_get_local_data=True,
        )
        detokenize_customer = detokenize_customers[0]
        first_payment = loan.payment_set.get(payment_number=1)
        return BNIMappingServices(
            loan=loan,
            first_payment=first_payment,
            customer=loan.customer,
            detokenize_customer=detokenize_customer,
        ).data_mapping(data_map=BNIDisbursementConst.DISBURSEMENT_DATA_MAPPING)

    @staticmethod
    def get_bni_supporting_disbursement_documents(
        loan: Loan, is_skip_download_files: bool
    ) -> Union[Dict[str, bytes], Dict[str, str]]:
        application = loan.get_application

        skrtp = Document.objects.filter(
            loan_xid=loan.loan_xid,
            document_source=loan.id,
            document_type__in=BNISupportingDisbursementDocumentConst.LIST_SKRTP_DOCUMENT_TYPE,
        ).last()
        if not skrtp:
            raise BNIChannelingLoanSKRTPNotFound()

        ktp = Image.objects.filter(image_source=application.id, image_type='ktp_self').last()
        if not ktp:
            raise BNIChannelingLoanKTPNotFound()

        selfie = Image.objects.filter(image_source=application.id, image_type='selfie').last()
        if not selfie:
            raise BNIChannelingLoanSelfieNotFound()

        if is_skip_download_files:
            return {
                BNISupportingDisbursementDocumentConst.SKRTP: skrtp.url,
                BNISupportingDisbursementDocumentConst.KTP: ktp.url,
                BNISupportingDisbursementDocumentConst.SELFIE: selfie.url,
            }

        return {
            BNISupportingDisbursementDocumentConst.SKRTP: download_file_from_oss(
                remote_filepath=skrtp.url
            ),
            BNISupportingDisbursementDocumentConst.KTP: download_file_from_oss(
                remote_filepath=ktp.url
            ),
            BNISupportingDisbursementDocumentConst.SELFIE: download_file_from_oss(
                remote_filepath=selfie.url
            ),
        }

    def get_list_bni_pending_channeling_loan_status(self) -> List[ChannelingLoanStatus]:
        return (
            ChannelingLoanStatus.objects.filter(
                channeling_type=self.channeling_type,
                channeling_status=ChannelingStatusConst.PENDING,
                channeling_eligibility_status__channeling_type=self.channeling_type,
                channeling_eligibility_status__eligibility_status=ChannelingStatusConst.ELIGIBLE,
                cdate__lte=self.current_ts,
            )
            .select_related('loan', 'loan__customer')
            .order_by("-pk")
        )

    def get_list_bni_recap_channeling_loan_status(self) -> Iterator[ChannelingLoanStatus]:
        return (
            ChannelingLoanStatus.objects.filter(
                channeling_type=self.channeling_type,
                channeling_status=ChannelingStatusConst.SUCCESS,
                cdate__gt=self.current_ts - timezone.timedelta(days=1),
                cdate__lte=self.current_ts,
            )
            .select_related('loan', 'loan__customer')
            .order_by("-pk")
            .iterator()  # data of whole day is quite big
        )

    def construct_list_bni_disbursement_data_and_supporting_documents(
        self,
        channeling_loan_statuses: Union[List[ChannelingLoanStatus], Iterator[ChannelingLoanStatus]],
        is_skip_download_supporting_docs: bool = False,
    ) -> Tuple[List[ChannelingLoanStatus], List[Dict[str, Any]], List[Dict[str, bytes]]]:
        list_disbursement_data = []
        list_supporting_documents = []
        success_construct_channeling_loan_statuses = []
        for channeling_loan_status in channeling_loan_statuses:
            try:
                loan = channeling_loan_status.loan
                disbursement_data = self.construct_bni_disbursement_data(loan=loan)
                supporting_documents = self.get_bni_supporting_disbursement_documents(
                    loan=loan, is_skip_download_files=is_skip_download_supporting_docs
                )
            except Exception as e:
                sentry_client.captureException()
                logger.exception(
                    {
                        'module': self.module,
                        'action': '{}.construct_list_bni_disbursement_data_'
                        'and_supporting_documents'.format(self.prefix_action),
                        'loan_id': channeling_loan_status.loan_id,
                        'message': 'Failed to construct channeling disbursement data for loan',
                        'error': e,
                    }
                )
                continue  # skip to the next loan

            success_construct_channeling_loan_statuses.append(channeling_loan_status)
            list_disbursement_data.append(disbursement_data)
            list_supporting_documents.append(supporting_documents)

        return (
            success_construct_channeling_loan_statuses,
            list_disbursement_data,
            list_supporting_documents,
        )

    @staticmethod
    def get_bni_disbursement_xlsx_headers() -> List[str]:
        headers = list(BNIDisbursementConst.DISBURSEMENT_DATA_MAPPING.keys())
        anchor_index = headers.index(BNIDisbursementConst.HEADER_ALAMAT_KANTOR_4)
        headers.insert(anchor_index + 1, BNIDisbursementConst.HEADER_KOTA_AREA)
        headers.insert(anchor_index + 1, BNIDisbursementConst.HEADER_KOTA_POS)
        headers.insert(anchor_index + 1, BNIDisbursementConst.HEADER_KOTA)
        return headers

    def construct_bni_disbursement_xlsx_bytes(
        self, list_disbursement_data: List[Dict[str, Any]]
    ) -> bytes:
        # Add duplicated header columns
        headers = self.get_bni_disbursement_xlsx_headers()
        return construct_bni_xlsx_bytes(
            headers=headers,
            sheet_name=BNIDisbursementConst.SHEET_NAME,
            list_data=list_disbursement_data,
            header_no=BNIDisbursementConst.HEADER_NO,
            header_map=None,
        )

    def construct_xlsx_file_and_upload_to_sftp_server(
        self,
        list_disbursement_data: List[Dict[str, Any]],
        filename_counter_suffix: str,
        is_recap: bool,
    ) -> str:
        # construct xlsx file name & content, then upload to bni sftp server
        xlsx_bytes = self.construct_bni_disbursement_xlsx_bytes(
            list_disbursement_data=list_disbursement_data
        )
        xlsx_filename = send_file_for_channeling_to_bni(
            data_bytes=xlsx_bytes,
            folder_name=BNIDisbursementConst.FOLDER_NAME,
            filename_format=BNIDisbursementConst.FILENAME_FORMAT,
            filename_datetime_format=(
                BNIDisbursementConst.FILENAME_DATETIME_FORMAT
                if not is_recap
                else BNIDisbursementConst.FILENAME_RECAP_DATETIME_FORMAT
            ),
            filename_counter_suffix=filename_counter_suffix,
            current_ts=self.current_ts,
        )
        return xlsx_filename

    @staticmethod
    def construct_bni_supporting_documents_zip_bytes(
        list_supporting_documents: List[Dict[str, bytes]]
    ) -> bytes:
        """
        Besides sending disbursement.xlsx that contains loan data,
        we also need to send the zip file that contains supporting documents.
        Each loan has 3 supporting documents: skrtp, ktp, and selfie
        """
        # Create a BytesIO object to hold the zip file in memory
        zip_buffer = BytesIO()

        # Create a ZipFile object
        with zipfile.ZipFile(zip_buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as zip_file:
            for index, supporting_documents in enumerate(list_supporting_documents):
                for document_type, document_data in supporting_documents.items():
                    filename = BNISupportingDisbursementDocumentConst.MAPPING_FILENAME_FORMAT[
                        document_type
                    ].format(index + 1)

                    # Write each file to the zip
                    zip_file.writestr(filename, document_data)

        # Seek to the beginning of the BytesIO buffer
        zip_buffer.seek(0)

        # Return the zip file as bytes
        return zip_buffer.getvalue()

    def construct_zip_file_and_upload_to_sftp_server(
        self,
        list_supporting_documents: List[Dict[str, bytes]],
        filename_counter_suffix: str,
        is_recap: bool,
    ) -> str:
        # construct zip file name & content, then upload to bni sftp server
        zip_bytes = self.construct_bni_supporting_documents_zip_bytes(
            list_supporting_documents=list_supporting_documents
        )
        zip_filename = send_file_for_channeling_to_bni(
            data_bytes=zip_bytes,
            folder_name=BNISupportingDisbursementDocumentConst.FOLDER_NAME,
            filename_format=BNISupportingDisbursementDocumentConst.FILENAME_FORMAT,
            filename_datetime_format=(
                BNISupportingDisbursementDocumentConst.FILENAME_DATETIME_FORMAT
                if not is_recap
                else BNISupportingDisbursementDocumentConst.FILENAME_RECAP_DATETIME_FORMAT
            ),
            filename_counter_suffix=filename_counter_suffix,
            current_ts=self.current_ts,
        )
        return zip_filename

    def approve_channeling_loan_statuses(
        self, channeling_loan_statuses: List[ChannelingLoanStatus]
    ):
        from juloserver.channeling_loan.tasks import approve_loan_for_channeling_task

        for channeling_loan_status in channeling_loan_statuses:
            approve_loan_for_channeling_task.delay(
                loan_id=channeling_loan_status.loan_id,
                channeling_type=self.channeling_type,
                approval_status=ChannelingApprovalStatusConst.YES,
            )

    def is_eligible_to_send_loan_to_bni(self) -> bool:
        channeling_loan_config = get_channeling_loan_configuration(
            channeling_type=self.channeling_type
        )

        # don't allow to run if feature setting is not active
        if not channeling_loan_config:
            return False

        cutoff_config = channeling_loan_config['cutoff']

        # allow to run if cutoff config is not active
        if not cutoff_config['is_active']:
            return True

        # don't allow to run if today exists in inactive day list
        if (
            convert_datetime_to_string(
                dt=self.current_ts, str_format=ChannelingConst.INACTIVE_DAY_STRING_FORMAT
            )
            in cutoff_config['INACTIVE_DAY']
        ):
            return False

        # don't allow to run if today exists in inactive date list
        if (
            convert_datetime_to_string(
                dt=self.current_ts, str_format=ChannelingConst.INACTIVE_DATE_STRING_FORMAT
            )
            in cutoff_config['INACTIVE_DATE']
        ):
            return False

        # allow to run if cutoff config is active and channel after cutoff flag is active
        if cutoff_config['CHANNEL_AFTER_CUTOFF']:
            return True

        # don't allow to run if current time is not in range of opening time and cutoff time
        opening_time = self.current_ts.replace(**cutoff_config['OPENING_TIME'])
        cutoff_time = self.current_ts.replace(**cutoff_config['CUTOFF_TIME'])
        if not (opening_time <= self.current_ts <= cutoff_time):
            return False

        return True

    def send_loan_for_channeling_to_bni(self) -> None:
        action_type = ChannelingActionTypeConst.DISBURSEMENT

        # CHECK EXECUTE OR NOT DEPENDS ON FEATURE SETTING
        if not self.is_eligible_to_send_loan_to_bni():
            logger.info(
                {
                    'module': self.module,
                    'action': '{}.send_loan_for_channeling_to_bni'.format(self.prefix_action),
                    'message': 'Skip to send loan for channeling to BNI due to feature setting',
                }
            )
            return None

        # CONSTRUCT DATA AND ONLY RETURN SUCCESS CONSTRUCT CHANNELING LOAN STATUSES
        # Disbursement doesn't need to send supported documents, only recap need it
        # However we still call the function to get the list of supporting documents to
        # skip loans that don't have enough documents but skip download files for better performance
        result_construct_data = self.construct_list_bni_disbursement_data_and_supporting_documents(
            channeling_loan_statuses=self.get_list_bni_pending_channeling_loan_status(),
            is_skip_download_supporting_docs=True,
        )

        # SUCCESS CONSTRUCT CHANNELING LOAN STATUSES, FAILED ALREADY BE SKIPPED
        channeling_loan_statuses = result_construct_data[0]
        if not channeling_loan_statuses:
            logger.info(
                {
                    'module': self.module,
                    'action': '{}.send_loan_for_channeling_to_bni'.format(self.prefix_action),
                    'message': 'No channeling loan statuses to send',
                }
            )
            return None

        filename_counter_suffix = get_next_filename_counter_suffix(
            channeling_type=self.channeling_type,
            action_type=action_type,
            current_ts=self.current_ts,
        )

        # CONSTRUCT AND UPLOAD XLSX FILE TO SFTP SERVER
        xlsx_filename = self.construct_xlsx_file_and_upload_to_sftp_server(
            list_disbursement_data=result_construct_data[1],
            filename_counter_suffix=filename_counter_suffix,
            is_recap=False,
        )

        # UPDATE CHANNELING LOAN STATUS FROM PENDING TO PROCESS
        bulk_update_channeling_loan_status(
            channeling_loan_status_id=[
                channeling_loan_status.id for channeling_loan_status in channeling_loan_statuses
            ],
            new_status=ChannelingStatusConst.PROCESS,
        )

        # PROCESS APPROVAL CHANNELING LOAN => CALL ASYNC TASK
        self.approve_channeling_loan_statuses(channeling_loan_statuses=channeling_loan_statuses)

        # CREATE A RECORD FOR TRACKING AND COUNTING FILE NAME
        create_channeling_loan_send_file_tracking(
            channeling_type=self.channeling_type,
            action_type=action_type,
        )

        logger.info(
            {
                'module': self.module,
                'action': '{}.send_loan_for_channeling_to_bni'.format(self.prefix_action),
                'number_of_loans': len(channeling_loan_statuses),
                'xlsx_filename': xlsx_filename,
                'message': 'Send loan for channeling to BNI successfully',
            }
        )

    def send_recap_loan_for_channeling_to_bni(self) -> None:
        action_type = ChannelingActionTypeConst.RECAP
        is_recap = True

        # CONSTRUCT DATA AND ONLY RETURN SUCCESS CONSTRUCT CHANNELING LOAN STATUSES
        result_construct_data = self.construct_list_bni_disbursement_data_and_supporting_documents(
            channeling_loan_statuses=self.get_list_bni_recap_channeling_loan_status(),
        )

        # SUCCESS CONSTRUCT CHANNELING LOAN STATUSES, FAILED ALREADY BE SKIPPED
        channeling_loan_statuses = result_construct_data[0]
        if not channeling_loan_statuses:
            logger.info(
                {
                    'module': self.module,
                    'action': '{}.send_recap_loan_for_channeling_to_bni'.format(self.prefix_action),
                    'message': 'No channeling loan statuses to send',
                }
            )
            return None

        filename_counter_suffix = get_next_filename_counter_suffix(
            channeling_type=self.channeling_type,
            action_type=action_type,
            current_ts=self.current_ts,
        )

        # CONSTRUCT AND UPLOAD XLSX FILE TO SFTP SERVER
        xlsx_filename = self.construct_xlsx_file_and_upload_to_sftp_server(
            list_disbursement_data=result_construct_data[1],
            filename_counter_suffix=filename_counter_suffix,
            is_recap=is_recap,
        )

        # CONSTRUCT AND UPLOAD ZIP FILE TO SFTP SERVER
        zip_filename = self.construct_zip_file_and_upload_to_sftp_server(
            list_supporting_documents=result_construct_data[2],
            filename_counter_suffix=filename_counter_suffix,
            is_recap=is_recap,
        )

        # CREATE A RECORD FOR TRACKING AND COUNTING FILE NAME
        create_channeling_loan_send_file_tracking(
            channeling_type=self.channeling_type,
            action_type=action_type,
        )

        logger.info(
            {
                'module': self.module,
                'action': '{}.send_recap_loan_for_channeling_to_bni'.format(self.prefix_action),
                'number_of_loans': len(channeling_loan_statuses),
                'xlsx_filename': xlsx_filename,
                'zip_filename': zip_filename,
                'message': 'Send recap loan for channeling to BNI successfully',
            }
        )


class BNIRepaymentServices:
    def __init__(self):
        self.action_type = ChannelingActionTypeConst.REPAYMENT
        self.channeling_type = ChannelingConst.BNI
        self.current_ts = timezone.localtime(timezone.now())
        self.module = 'juloserver.channeling_loan'
        self.prefix_action = 'services.bni_services.BNIRepaymentServices'

    def get_bni_repayment_data(self, request) -> List[Dict[str, Any]]:
        """
        Convert excel to List[Dict[str, Any]] format so can reuse main function
        capture error in case conversion fail
        """
        upload_form = RepaymentFileForm(request.POST, request.FILES)
        if not upload_form.is_valid():
            raise Exception('Invalid form')

        file_ = upload_form.cleaned_data['repayment_file_field']
        extension = file_.name.split('.')[-1]

        if extension not in ChannelingConst.FILE_UPLOAD_EXTENSIONS:
            raise Exception('Please upload correct file excel')

        try:
            excel_datas = xls_to_dict(file_)
            list_data = []
            for sheet in excel_datas:
                for _, row in enumerate(excel_datas[sheet]):
                    list_data.append(row)

            return list_data
        except Exception as error:
            sentry_client.captureException()
            logger.exception(
                {
                    'module': self.module,
                    'action': '{}.get_bni_repayment_data'.format(self.prefix_action),
                    'message': 'Failed to construct channeling repayment data',
                    'error': error,
                }
            )
            raise Exception(error)

    def construct_bni_repayment_xlsx_bytes(
        self, list_repayment_data: List[Dict[str, Any]]
    ) -> bytes:
        """
        return excel content on bytes
        """
        headers = BNIRepaymentConst.REPAYMENT_EXTRA_HEADER
        if list_repayment_data:
            headers = headers + list(list_repayment_data[0].keys())

        return construct_bni_xlsx_bytes(
            headers=headers,
            sheet_name=BNIRepaymentConst.SHEET_NAME,
            list_data=list_repayment_data,
            header_no=BNIRepaymentConst.HEADER_NO,
            header_map=BNIRepaymentConst.REPAYMENT_HEADER_DATA_MAPPING,
        )

    def send_repayment_for_channeling_to_bni(self, request) -> Tuple[bool, Dict[str, Any]]:
        """
        main function to process the form request
        form are validated, then sent to BNI SFTP
        """
        try:
            list_repayment_data = self.get_bni_repayment_data(request)
        except Exception as error:
            return (
                False,
                {
                    'content': str(error),
                    'filename': None,
                },
            )

        xlsx_bytes = self.construct_bni_repayment_xlsx_bytes(
            list_repayment_data=list_repayment_data
        )

        filename_counter_suffix = get_next_filename_counter_suffix(
            channeling_type=self.channeling_type,
            action_type=self.action_type,
            current_ts=self.current_ts,
        )
        # SEND EXCEL FILE TO BNI SFTP
        filename = send_file_for_channeling_to_bni(
            data_bytes=xlsx_bytes,
            folder_name=BNIRepaymentConst.FOLDER_NAME,
            filename_format=BNIRepaymentConst.FILENAME_FORMAT,
            filename_datetime_format=BNIRepaymentConst.FILENAME_DATETIME_FORMAT,
            filename_counter_suffix=filename_counter_suffix,
            current_ts=self.current_ts,
        )

        # CREATE A RECORD FOR TRACKING AND COUNTING FILE NAME
        create_channeling_loan_send_file_tracking(
            channeling_type=self.channeling_type,
            action_type=self.action_type,
        )
        logger.info(
            {
                'module': self.module,
                'action': '{}.send_repayment_for_channeling_to_bni'.format(self.prefix_action),
                'number_of_payments': len(list_repayment_data),
                'filename': filename,
                'message': 'Send loan for channeling to BNI successfully',
            }
        )
        return True, {
            'content': xlsx_bytes,
            'filename': filename,
        }
